"""Excel/CSV bulk import — ingest raw Ninehire export or CSV and create candidate entries."""
import csv
import io
import os
import uuid
import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from backend.config import UPLOAD_DIR
from backend.database import get_db
from backend.models import Candidate, Document

router = APIRouter(prefix="/api/import", tags=["import"])

NAME_KEYWORDS = ["회사", "브랜드", "기업", "업체", "상호", "이름", "name", "company"]


def _find_name_col(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if any(kw in h.lower() for kw in NAME_KEYWORDS):
            return i
    return 0


def _import_rows(headers: list[str], rows: list, filename: str, file_path: str, file_type: str, db: Session) -> list[dict]:
    name_col = _find_name_col(headers)
    imported = []

    for row in rows:
        if not row or not any(row):
            continue

        name = str(row[name_col]).strip() if row[name_col] else "이름 없음"
        if name in ("None", ""):
            name = "이름 없음"

        row_data = {}
        for i, val in enumerate(row):
            if i < len(headers) and val is not None and str(val).strip():
                row_data[headers[i]] = str(val).strip()

        extracted_text = "\n".join(f"{k}: {v}" for k, v in row_data.items())

        existing = db.query(Candidate).filter(Candidate.name == name).first()
        if existing:
            candidate = existing
        else:
            candidate = Candidate(name=name)
            db.add(candidate)
            db.commit()
            db.refresh(candidate)

        doc = Document(
            candidate_id=candidate.id,
            filename=filename,
            file_path=file_path,
            file_type=file_type,
            extracted_text=extracted_text,
        )
        db.add(doc)
        db.commit()

        imported.append({"id": candidate.id, "name": name, "data": row_data})

    return imported


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import candidates from Excel or CSV."""
    ext = os.path.splitext(file.filename or "")[1].lower().lstrip(".")
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx/.xls) 또는 CSV(.csv) 파일만 지원합니다")

    unique_name = f"{uuid.uuid4().hex}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, unique_name)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    if ext == "csv":
        headers, rows = _parse_csv(content)
        file_type = "csv"
    else:
        headers, rows = _parse_excel(file_path)
        file_type = "xlsx"

    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="데이터가 없습니다")

    imported = _import_rows(headers, rows, file.filename or f"import.{ext}", file_path, file_type, db)

    return {
        "message": f"{len(imported)}개 후보 가져오기 완료",
        "count": len(imported),
        "candidates": imported,
        "headers": headers,
    }


class LocalImport(BaseModel):
    file_path: str


@router.post("/local")
def import_local_file(data: LocalImport, db: Session = Depends(get_db)):
    """Import from a local file path (called from Claude Code skill)."""
    path = data.file_path.replace("\\", "/")
    if not os.path.exists(path):
        raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {path}")

    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx/.xls) 또는 CSV(.csv) 파일만 지원합니다")

    filename = os.path.basename(path)

    if ext == "csv":
        with open(path, "rb") as f:
            content = f.read()
        headers, rows = _parse_csv(content)
        file_type = "csv"
    else:
        headers, rows = _parse_excel(path)
        file_type = "xlsx"

    if len(rows) < 1:
        raise HTTPException(status_code=400, detail="데이터가 없습니다")

    imported = _import_rows(headers, rows, filename, path, file_type, db)

    return {
        "message": f"{len(imported)}개 후보 가져오기 완료",
        "count": len(imported),
        "candidates": imported,
        "headers": headers,
    }


def _parse_csv(content: bytes) -> tuple[list[str], list]:
    """Parse CSV bytes, auto-detect encoding and delimiter."""
    for encoding in ("utf-8-sig", "utf-8", "euc-kr", "cp949"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = content.decode("utf-8", errors="replace")

    # Auto-detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:2000])
    except csv.Error:
        dialect = None

    reader = csv.reader(io.StringIO(text), dialect or csv.excel)
    all_rows = list(reader)
    if len(all_rows) < 2:
        return [], []

    headers = [h.strip() if h else f"열{i}" for i, h in enumerate(all_rows[0])]
    data_rows = all_rows[1:]
    return headers, data_rows


def _parse_excel(file_path: str) -> tuple[list[str], list]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        return [], []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], []

    headers = [str(h).strip() if h else f"열{i}" for i, h in enumerate(rows[0])]
    return headers, rows[1:]


@router.post("/excel/preview", response_class=HTMLResponse)
async def preview_excel(file: UploadFile = File(...)):
    """Preview Excel data as clean HTML table without importing."""
    ext = os.path.splitext(file.filename or "")[1].lower().lstrip(".")
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx/.xls)만 지원합니다")

    content = await file.read()
    unique_name = f"preview_{uuid.uuid4().hex}.xlsx"
    tmp_path = os.path.join(UPLOAD_DIR, unique_name)
    with open(tmp_path, "wb") as f:
        f.write(content)

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    os.remove(tmp_path)

    if len(rows) < 1:
        return "<p>데이터가 없습니다</p>"

    headers = [str(h).strip() if h else "" for h in rows[0]]

    html = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<style>
  body { font-family: 'Pretendard', -apple-system, sans-serif; margin: 2rem; background: #f8f9fa; }
  h1 { color: #1a1a2e; font-size: 1.5rem; }
  .info { color: #666; margin-bottom: 1rem; }
  table { border-collapse: collapse; width: 100%; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
  th { background: #1a1a2e; color: white; padding: 12px 16px; text-align: left; font-size: 0.85rem; white-space: nowrap; }
  td { padding: 10px 16px; border-bottom: 1px solid #eee; font-size: 0.85rem; max-width: 300px; overflow: hidden; text-overflow: ellipsis; }
  tr:hover td { background: #f0f4ff; }
  tr:nth-child(even) td { background: #fafafa; }
  tr:nth-child(even):hover td { background: #f0f4ff; }
  .count { background: #e8f0fe; color: #1a73e8; padding: 4px 12px; border-radius: 12px; font-size: 0.85rem; }
</style>
</head>
<body>
"""
    html += f"<h1>📋 지원자 데이터 미리보기</h1>"
    html += f'<p class="info">파일: {file.filename} | <span class="count">총 {len(rows) - 1}건</span></p>'
    html += "<table><thead><tr>"
    for h in headers:
        html += f"<th>{h}</th>"
    html += "</tr></thead><tbody>"

    for row in rows[1:]:
        html += "<tr>"
        for val in row:
            display = str(val).strip() if val is not None else ""
            if len(display) > 100:
                display = display[:100] + "..."
            html += f"<td>{display}</td>"
        html += "</tr>"

    html += "</tbody></table></body></html>"
    return html
